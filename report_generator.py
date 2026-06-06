import os
from datetime import datetime, timedelta
from typing import Dict, Any, List, Optional, Tuple

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import pandas as pd
import numpy as np

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm, cm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    Image, PageBreak, HRFlowable
)
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.drawing.image import Image as XLImage

from config import REPORT_DIR, SUPPLIER_LEVELS
from models import (
    SessionLocal, Supplier, Sample, InspectionTask,
    RectificationOrder, MonthlyReport
)
from notification import log_operation, push_alert, write_audit_log


plt.rcParams["font.sans-serif"] = ["SimHei", "Microsoft YaHei", "DejaVu Sans"]
plt.rcParams["axes.unicode_minus"] = False


def _register_chinese_font():
    font_paths = [
        "C:/Windows/Fonts/msyh.ttc",
        "C:/Windows/Fonts/simhei.ttf",
        "C:/Windows/Fonts/simsun.ttc",
        "/usr/share/fonts/truetype/wqy/wqy-microhei.ttc",
        "/System/Library/Fonts/PingFang.ttc",
    ]
    for fp in font_paths:
        if os.path.exists(fp):
            try:
                pdfmetrics.registerFont(TTFont("ChineseFont", fp))
                return "ChineseFont"
            except Exception:
                continue
    return "Helvetica"


CHINESE_FONT = _register_chinese_font()


def get_month_range(year: int, month: int) -> Tuple[datetime, datetime]:
    start = datetime(year, month, 1)
    if month == 12:
        end = datetime(year + 1, 1, 1)
    else:
        end = datetime(year, month + 1, 1)
    return start, end


def calculate_monthly_stats(year: int, month: int) -> Dict[str, Any]:
    db = SessionLocal()
    try:
        start_date, end_date = get_month_range(year, month)

        samples = db.query(Sample).filter(
            Sample.received_at >= start_date,
            Sample.received_at < end_date
        ).all()

        supplier_stats = {}
        total_pass = 0
        total_count = 0
        total_cycle_days = 0.0
        cycle_count = 0

        for sample in samples:
            supplier_id = sample.supplier_id
            if supplier_id not in supplier_stats:
                sup = db.query(Supplier).filter(Supplier.id == supplier_id).first()
                supplier_stats[supplier_id] = {
                    "supplier_code": sup.supplier_code if sup else "UNKNOWN",
                    "supplier_name": sup.supplier_name if sup else "未知",
                    "level": sup.level if sup else "B",
                    "total": 0,
                    "pass": 0,
                    "fail": 0,
                    "pending": 0,
                    "cycle_days_list": []
                }

            s = supplier_stats[supplier_id]
            s["total"] += 1
            total_count += 1

            if sample.status == "合格":
                s["pass"] += 1
                total_pass += 1
            elif sample.status == "不合格":
                s["fail"] += 1
            else:
                s["pending"] += 1

            tasks = db.query(InspectionTask).filter(
                InspectionTask.sample_id == sample.id
            ).all()
            if tasks and sample.received_at:
                completed = [t for t in tasks if t.inspected_at]
                if completed:
                    last_inspected = max(t.inspected_at for t in completed)
                    cycle = (last_inspected - sample.received_at).total_seconds() / 86400
                    s["cycle_days_list"].append(cycle)
                    total_cycle_days += cycle
                    cycle_count += 1

        result_list = []
        for sid, stats in supplier_stats.items():
            pass_rate = (stats["pass"] / stats["total"] * 100) if stats["total"] > 0 else 0.0
            avg_cycle = (
                sum(stats["cycle_days_list"]) / len(stats["cycle_days_list"])
                if stats["cycle_days_list"] else 0.0
            )
            result_list.append({
                "supplier_id": sid,
                "supplier_code": stats["supplier_code"],
                "supplier_name": stats["supplier_name"],
                "level": stats["level"],
                "level_desc": SUPPLIER_LEVELS.get(stats["level"], {}).get("description", ""),
                "total": stats["total"],
                "pass": stats["pass"],
                "fail": stats["fail"],
                "pending": stats["pending"],
                "pass_rate": round(pass_rate, 2),
                "avg_cycle_days": round(avg_cycle, 2)
            })

        result_list.sort(key=lambda x: (-x["pass_rate"], x["avg_cycle_days"]))
        for idx, item in enumerate(result_list, 1):
            item["ranking"] = idx

        rectify_count = db.query(RectificationOrder).filter(
            RectificationOrder.created_at >= start_date,
            RectificationOrder.created_at < end_date
        ).count()

        overall_pass_rate = (total_pass / total_count * 100) if total_count > 0 else 0.0
        overall_avg_cycle = (total_cycle_days / cycle_count) if cycle_count > 0 else 0.0

        return {
            "year": year,
            "month": month,
            "period": f"{year}年{month}月",
            "total_samples": total_count,
            "total_pass": total_pass,
            "total_fail": total_count - total_pass - sum(s["pending"] for s in result_list),
            "pending": sum(s["pending"] for s in result_list),
            "overall_pass_rate": round(overall_pass_rate, 2),
            "overall_avg_cycle_days": round(overall_avg_cycle, 2),
            "rectification_count": rectify_count,
            "suppliers": result_list,
            "supplier_count": len(result_list)
        }
    finally:
        db.close()


def get_trend_data(months: int = 6) -> Dict[str, Any]:
    db = SessionLocal()
    try:
        today = datetime.now()
        trend = []
        for i in range(months - 1, -1, -1):
            y = today.year
            m = today.month - i
            if m <= 0:
                m += 12
                y -= 1
            stats = calculate_monthly_stats(y, m)
            trend.append({
                "period": f"{y}-{m:02d}",
                "year": y,
                "month": m,
                "total_samples": stats["total_samples"],
                "pass_rate": stats["overall_pass_rate"],
                "avg_cycle_days": stats["overall_avg_cycle_days"]
            })
        return {"months": months, "trend": trend}
    finally:
        db.close()


def generate_trend_charts(trend_data: Dict[str, Any], output_dir: str) -> Dict[str, str]:
    trend = trend_data["trend"]
    periods = [t["period"] for t in trend]
    pass_rates = [t["pass_rate"] for t in trend]
    sample_counts = [t["total_samples"] for t in trend]
    avg_cycles = [t["avg_cycle_days"] for t in trend]

    chart_paths = {}

    fig, ax1 = plt.subplots(figsize=(10, 5))
    x = np.arange(len(periods))
    width = 0.35
    bars = ax1.bar(x - width / 2, sample_counts, width, label="检测数量", color="#4A90D9", alpha=0.8)
    ax1.set_xlabel("月份")
    ax1.set_ylabel("检测数量", color="#4A90D9")
    ax1.tick_params(axis="y", labelcolor="#4A90D9")
    ax1.set_xticks(x)
    ax1.set_xticklabels(periods, rotation=45)

    ax2 = ax1.twinx()
    line = ax2.plot(x, pass_rates, "o-", color="#E74C3C", linewidth=2, markersize=8, label="合格率(%)")
    ax2.set_ylabel("合格率 (%)", color="#E74C3C")
    ax2.tick_params(axis="y", labelcolor="#E74C3C")
    ax2.set_ylim(0, 105)

    for bar, val in zip(bars, sample_counts):
        ax1.text(bar.get_x() + bar.get_width() / 2, bar.get_height(),
                 str(val), ha="center", va="bottom", fontsize=10)
    for i, val in enumerate(pass_rates):
        ax2.text(i, val + 1, f"{val}%", ha="center", va="bottom",
                 fontsize=10, color="#E74C3C", fontweight="bold")

    lines1, labels1 = ax1.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    ax1.legend(lines1 + lines2, labels1 + labels2, loc="lower right")

    plt.title(f"近{trend_data['months']}个月检测数量与合格率趋势", fontsize=14, fontweight="bold")
    plt.tight_layout()
    p1 = os.path.join(output_dir, "trend_pass_rate.png")
    plt.savefig(p1, dpi=150, bbox_inches="tight")
    plt.close()
    chart_paths["pass_rate"] = p1

    fig, ax = plt.subplots(figsize=(10, 5))
    bars = ax.bar(periods, avg_cycles, color="#27AE60", alpha=0.8, width=0.5)
    ax.set_xlabel("月份")
    ax.set_ylabel("平均检测周期 (天)")
    ax.set_title(f"近{trend_data['months']}个月平均检测周期趋势", fontsize=14, fontweight="bold")
    ax.set_xticklabels(periods, rotation=45)

    for bar, val in zip(bars, avg_cycles):
        ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height(),
                f"{val:.1f}天", ha="center", va="bottom", fontsize=10, fontweight="bold")

    avg_line = np.mean(avg_cycles) if avg_cycles else 0
    ax.axhline(y=avg_line, color="#E67E22", linestyle="--", linewidth=2, label=f"平均: {avg_line:.1f}天")
    ax.legend()

    plt.tight_layout()
    p2 = os.path.join(output_dir, "trend_cycle.png")
    plt.savefig(p2, dpi=150, bbox_inches="tight")
    plt.close()
    chart_paths["cycle"] = p2

    return chart_paths


def generate_ranking_chart(stats: Dict[str, Any], output_dir: str) -> str:
    suppliers = stats["suppliers"][:15]
    names = [s["supplier_name"][:10] for s in suppliers]
    pass_rates = [s["pass_rate"] for s in suppliers]
    colors_list = ["#27AE60" if r >= 95 else "#F39C12" if r >= 80 else "#E74C3C" for r in pass_rates]

    fig, ax = plt.subplots(figsize=(12, 6))
    y_pos = np.arange(len(names))
    bars = ax.barh(y_pos, pass_rates, color=colors_list, alpha=0.85, height=0.6)
    ax.set_yticks(y_pos)
    ax.set_yticklabels(names)
    ax.invert_yaxis()
    ax.set_xlabel("合格率 (%)")
    ax.set_title(f"{stats['period']} 供应商质量合格率排名", fontsize=14, fontweight="bold")
    ax.set_xlim(0, 105)
    ax.axvline(x=90, color="#2980B9", linestyle=":", linewidth=2, label="目标线 90%")

    for bar, val in zip(bars, pass_rates):
        ax.text(bar.get_width() + 1, bar.get_y() + bar.get_height() / 2,
                f"{val}%", ha="left", va="center", fontsize=10, fontweight="bold")

    ax.legend()
    plt.tight_layout()
    path = os.path.join(output_dir, "supplier_ranking.png")
    plt.savefig(path, dpi=150, bbox_inches="tight")
    plt.close()
    return path


def generate_pdf_report(stats: Dict[str, Any], chart_paths: Dict[str, str],
                        ranking_chart: str, output_path: str) -> str:
    doc = SimpleDocTemplate(
        output_path, pagesize=A4,
        rightMargin=20 * mm, leftMargin=20 * mm,
        topMargin=20 * mm, bottomMargin=20 * mm
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "CustomTitle", parent=styles["Title"],
        fontName=CHINESE_FONT, fontSize=22, leading=28,
        alignment=1, spaceAfter=20
    )
    h2_style = ParagraphStyle(
        "CustomH2", parent=styles["Heading2"],
        fontName=CHINESE_FONT, fontSize=14, leading=20,
        spaceBefore=15, spaceAfter=10
    )
    normal_style = ParagraphStyle(
        "CustomNormal", parent=styles["Normal"],
        fontName=CHINESE_FONT, fontSize=10, leading=16
    )

    story = []
    story.append(Paragraph(f"供应商样品检测质量月报", title_style))
    story.append(Paragraph(f"统计周期：{stats['period']}", ParagraphStyle(
        "SubTitle", parent=normal_style, alignment=1, fontSize=12, spaceAfter=30
    )))
    story.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor("#2980B9")))
    story.append(Spacer(1, 10))

    story.append(Paragraph("一、总体概况", h2_style))
    overview_data = [
        ["指标", "数值"],
        ["检测样品总数", str(stats["total_samples"])],
        ["合格数量", str(stats["total_pass"])],
        ["不合格数量", str(stats["total_fail"])],
        ["待检数量", str(stats["pending"])],
        ["整体合格率", f"{stats['overall_pass_rate']}%"],
        ["平均检测周期", f"{stats['overall_avg_cycle_days']} 天"],
        ["整改工单数量", str(stats["rectification_count"])],
        ["参与供应商数", str(stats["supplier_count"])]
    ]
    t = Table(overview_data, colWidths=[100 * mm, 80 * mm])
    t.setStyle(TableStyle([
        ("FONT", (0, 0), (-1, -1), CHINESE_FONT, 10),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2980B9")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#ECF0F1")]),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(t)
    story.append(Spacer(1, 15))

    story.append(Paragraph("二、趋势分析", h2_style))
    if os.path.exists(chart_paths.get("pass_rate", "")):
        story.append(Image(chart_paths["pass_rate"], width=170 * mm, height=85 * mm))
        story.append(Spacer(1, 10))
    if os.path.exists(chart_paths.get("cycle", "")):
        story.append(Image(chart_paths["cycle"], width=170 * mm, height=85 * mm))

    story.append(PageBreak())
    story.append(Paragraph("三、供应商质量排名", h2_style))
    if os.path.exists(ranking_chart):
        story.append(Image(ranking_chart, width=170 * mm, height=90 * mm))
    story.append(Spacer(1, 10))

    header = ["排名", "供应商代码", "供应商名称", "等级", "送检数", "合格数", "合格率", "平均周期(天)"]
    table_data = [header]
    for s in stats["suppliers"]:
        table_data.append([
            str(s["ranking"]),
            s["supplier_code"],
            s["supplier_name"][:14],
            s["level"],
            str(s["total"]),
            str(s["pass"]),
            f"{s['pass_rate']}%",
            str(s["avg_cycle_days"])
        ])

    col_w = [12 * mm, 28 * mm, 50 * mm, 14 * mm, 18 * mm, 18 * mm, 22 * mm, 28 * mm]
    t2 = Table(table_data, colWidths=col_w, repeatRows=1)

    def _row_color(row_idx, _):
        if row_idx == 0:
            return colors.HexColor("#2980B9")
        try:
            rate = float(str(table_data[row_idx][6]).replace("%", ""))
        except (ValueError, IndexError):
            rate = 100
        if rate >= 95:
            return colors.HexColor("#D5F5E3")
        elif rate >= 80:
            return colors.HexColor("#FEF9E7")
        else:
            return colors.HexColor("#FADBD8")

    style_cmds = [
        ("FONT", (0, 0), (-1, -1), CHINESE_FONT, 9),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]
    for i in range(len(table_data)):
        style_cmds.append(("BACKGROUND", (0, i), (-1, i), _row_color(i, None)))
    t2.setStyle(TableStyle(style_cmds))
    story.append(t2)

    story.append(Spacer(1, 20))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.grey))
    story.append(Spacer(1, 8))
    story.append(Paragraph(
        f"报告生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}　｜　生成系统：供应商样品检测质量管理系统",
        ParagraphStyle("Footer", parent=normal_style, fontSize=9, textColor=colors.grey, alignment=1)
    ))

    doc.build(story)
    return output_path


def generate_excel_report(stats: Dict[str, Any], trend_data: Dict[str, Any],
                          chart_paths: Dict[str, str], ranking_chart: str,
                          output_path: str) -> str:
    wb = Workbook()

    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    header_fill = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws1 = wb.active
    ws1.title = "总体概况"

    ws1["A1"] = f"供应商样品检测质量月报 - {stats['period']}"
    ws1.merge_cells("A1:F1")
    c = ws1["A1"]
    c.font = Font(size=16, bold=True)
    c.alignment = center_align
    ws1.row_dimensions[1].height = 30

    overview = [
        ["指标", "数值", "", "指标", "数值"],
        ["检测样品总数", stats["total_samples"], "", "整体合格率", f"{stats['overall_pass_rate']}%"],
        ["合格数量", stats["total_pass"], "", "平均检测周期", f"{stats['overall_avg_cycle_days']} 天"],
        ["不合格数量", stats["total_fail"], "", "整改工单数量", stats["rectification_count"]],
        ["待检数量", stats["pending"], "", "参与供应商数", stats["supplier_count"]],
    ]
    for r_idx, row in enumerate(overview, 3):
        for c_idx, val in enumerate(row, 1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=val)
            cell.border = thin_border
            cell.alignment = center_align
            if r_idx == 3:
                cell.fill = header_fill
                cell.font = header_font

    ws1.column_dimensions["A"].width = 20
    ws1.column_dimensions["B"].width = 15
    ws1.column_dimensions["D"].width = 20
    ws1.column_dimensions["E"].width = 18

    ws2 = wb.create_sheet("供应商排名")
    headers2 = ["排名", "供应商代码", "供应商名称", "等级", "等级说明",
                "送检总数", "合格数", "不合格数", "待检数", "合格率(%)", "平均检测周期(天)"]
    for c_idx, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=c_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    for r_idx, s in enumerate(stats["suppliers"], 2):
        row_data = [
            s["ranking"], s["supplier_code"], s["supplier_name"], s["level"],
            s["level_desc"], s["total"], s["pass"], s["fail"], s["pending"],
            s["pass_rate"], s["avg_cycle_days"]
        ]
        for c_idx, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = center_align
            cell.border = thin_border
            try:
                rate = float(s["pass_rate"])
            except (ValueError, TypeError):
                rate = 100
            if rate >= 95:
                cell.fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
            elif rate >= 80:
                cell.fill = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")

    widths = [8, 14, 30, 8, 18, 10, 10, 10, 10, 12, 16]
    for i, w in enumerate(widths, 1):
        ws2.column_dimensions[chr(64 + i)].width = w
    ws2.freeze_panes = "A2"

    ws3 = wb.create_sheet("趋势数据")
    headers3 = ["月份", "检测数量", "合格率(%)", "平均检测周期(天)"]
    for c_idx, h in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=c_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    for r_idx, t in enumerate(trend_data["trend"], 2):
        row_data = [t["period"], t["total_samples"], t["pass_rate"], t["avg_cycle_days"]]
        for c_idx, val in enumerate(row_data, 1):
            cell = ws3.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = center_align
            cell.border = thin_border
    for i, w in enumerate([14, 12, 14, 18], 1):
        ws3.column_dimensions[chr(64 + i)].width = w

    try:
        if os.path.exists(chart_paths.get("pass_rate", "")):
            img = XLImage(chart_paths["pass_rate"])
            img.width = 600
            img.height = 300
            ws3.add_image(img, "F2")
        if os.path.exists(chart_paths.get("cycle", "")):
            img2 = XLImage(chart_paths["cycle"])
            img2.width = 600
            img2.height = 300
            ws3.add_image(img2, "F20")
    except Exception:
        pass

    try:
        if os.path.exists(ranking_chart):
            ws4 = wb.create_sheet("图表")
            img3 = XLImage(ranking_chart)
            img3.width = 700
            img3.height = 400
            ws4.add_image(img3, "A1")
    except Exception:
        pass

    wb.save(output_path)
    return output_path


def generate_monthly_report(year: Optional[int] = None,
                            month: Optional[int] = None,
                            operator: str = "system") -> Dict[str, Any]:
    now = datetime.now()
    if year is None:
        year = now.year
    if month is None:
        month = now.month

    report_label = f"{year}{month:02d}"
    pdf_path = os.path.join(REPORT_DIR, f"质量月报_{report_label}.pdf")
    excel_path = os.path.join(REPORT_DIR, f"质量月报_{report_label}.xlsx")

    try:
        stats = calculate_monthly_stats(year, month)
        trend_data = get_trend_data(6)
        chart_paths = generate_trend_charts(trend_data, REPORT_DIR)
        ranking_chart = generate_ranking_chart(stats, REPORT_DIR)

        generate_pdf_report(stats, chart_paths, ranking_chart, pdf_path)
        generate_excel_report(stats, trend_data, chart_paths, ranking_chart, excel_path)

        db = SessionLocal()
        try:
            existing = db.query(MonthlyReport).filter(
                MonthlyReport.report_year == year,
                MonthlyReport.report_month == month
            ).first()
            if existing:
                existing.pdf_path = pdf_path
                existing.excel_path = excel_path
                existing.total_samples = stats["total_samples"]
                existing.pass_rate = stats["overall_pass_rate"]
                existing.avg_cycle_days = stats["overall_avg_cycle_days"]
                existing.generated_at = datetime.now()
            else:
                mr = MonthlyReport(
                    report_month=f"{year}-{month:02d}",
                    report_year=year,
                    pdf_path=pdf_path,
                    excel_path=excel_path,
                    generated_by=operator,
                    total_samples=stats["total_samples"],
                    pass_rate=stats["overall_pass_rate"],
                    avg_cycle_days=stats["overall_avg_cycle_days"]
                )
                db.add(mr)
            db.commit()
        finally:
            db.close()

        log_operation(
            "生成月度报告",
            f"{year}年{month}月报告已生成：PDF={pdf_path}，Excel={excel_path}",
            operator=operator
        )
        write_audit_log(
            f"月度报告生成 | {year}年{month}月 | 样品数:{stats['total_samples']} | "
            f"合格率:{stats['overall_pass_rate']}%"
        )

        push_alert(
            "📊 月度质量报告已生成",
            f"**报告周期**: {year}年{month}月\n"
            f"**检测样品数**: {stats['total_samples']}\n"
            f"**整体合格率**: {stats['overall_pass_rate']}%\n"
            f"**平均检测周期**: {stats['overall_avg_cycle_days']} 天\n"
            f"**参与供应商**: {stats['supplier_count']} 家\n"
            f"**PDF报告**: {os.path.basename(pdf_path)}\n"
            f"**Excel报告**: {os.path.basename(excel_path)}",
            level="info"
        )

        return {
            "success": True,
            "period": f"{year}年{month}月",
            "pdf_path": pdf_path,
            "excel_path": excel_path,
            "stats": {
                "total_samples": stats["total_samples"],
                "pass_rate": stats["overall_pass_rate"],
                "avg_cycle_days": stats["overall_avg_cycle_days"],
                "supplier_count": stats["supplier_count"]
            }
        }
    except Exception as e:
        return {"success": False, "error": str(e)}
