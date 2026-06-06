import os
from datetime import datetime
from typing import Dict, Any, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from sqlalchemy import and_, or_

from config import EXPORT_DIR
from models import (
    SessionLocal, Sample, Supplier, InspectionTask,
    RectificationOrder, OperationLog
)
from notification import log_operation, write_audit_log


def _parse_date(date_str: Optional[str]) -> Optional[datetime]:
    if not date_str:
        return None
    for fmt in ["%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d"]:
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            continue
    return None


def query_samples(
    supplier_name: Optional[str] = None,
    supplier_code: Optional[str] = None,
    sample_type: Optional[str] = None,
    inspection_item: Optional[str] = None,
    status: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    sample_code: Optional[str] = None,
    product_name: Optional[str] = None,
    page: int = 1,
    page_size: int = 50
) -> Dict[str, Any]:
    db = SessionLocal()
    try:
        query = db.query(Sample).join(Supplier, Sample.supplier_id == Supplier.id)

        if supplier_name:
            query = query.filter(Supplier.supplier_name.like(f"%{supplier_name}%"))
        if supplier_code:
            query = query.filter(Supplier.supplier_code.like(f"%{supplier_code}%"))
        if sample_type:
            query = query.filter(Sample.sample_type.like(f"%{sample_type}%"))
        if status:
            query = query.filter(Sample.status == status)
        if sample_code:
            query = query.filter(Sample.sample_code.like(f"%{sample_code}%"))
        if product_name:
            query = query.filter(Sample.product_name.like(f"%{product_name}%"))

        start_dt = _parse_date(start_date)
        end_dt = _parse_date(end_date)
        if start_dt:
            query = query.filter(Sample.received_at >= start_dt)
        if end_dt:
            query = query.filter(Sample.received_at < end_dt)

        if inspection_item:
            sub = db.query(InspectionTask.sample_id).filter(
                InspectionTask.inspection_item.like(f"%{inspection_item}%")
            ).subquery()
            query = query.filter(Sample.id.in_(sub))

        total = query.count()
        query = query.order_by(Sample.received_at.desc())
        offset = (page - 1) * page_size
        samples = query.offset(offset).limit(page_size).all()

        results = []
        for s in samples:
            tasks = db.query(InspectionTask).filter(
                InspectionTask.sample_id == s.id
            ).all()
            pass_count = sum(1 for t in tasks if t.status == "合格")
            fail_count = sum(1 for t in tasks if t.status == "不合格")
            results.append({
                "sample_id": s.id,
                "sample_code": s.sample_code,
                "supplier_code": s.supplier.supplier_code if s.supplier else "",
                "supplier_name": s.supplier.supplier_name if s.supplier else "",
                "supplier_level": s.supplier.level if s.supplier else "",
                "sample_type": s.sample_type,
                "product_name": s.product_name,
                "product_spec": s.product_spec or "",
                "supplier_batch": s.supplier_batch or "",
                "quantity": s.quantity,
                "status": s.status,
                "received_at": s.received_at.strftime("%Y-%m-%d %H:%M:%S") if s.received_at else "",
                "received_by": s.received_by or "",
                "task_count": len(tasks),
                "pass_count": pass_count,
                "fail_count": fail_count,
                "inspection_items": [t.inspection_item for t in tasks],
                "remarks": s.remarks or ""
            })

        return {
            "success": True,
            "total": total,
            "page": page,
            "page_size": page_size,
            "total_pages": (total + page_size - 1) // page_size if page_size > 0 else 0,
            "data": results
        }
    finally:
        db.close()


def query_inspection_tasks(
    supplier_name: Optional[str] = None,
    inspection_item: Optional[str] = None,
    laboratory: Optional[str] = None,
    inspector: Optional[str] = None,
    status: Optional[str] = None,
    result_status: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    page: int = 1,
    page_size: int = 50
) -> Dict[str, Any]:
    db = SessionLocal()
    try:
        query = db.query(InspectionTask).join(Sample, InspectionTask.sample_id == Sample.id) \
                                        .join(Supplier, Sample.supplier_id == Supplier.id)

        if supplier_name:
            query = query.filter(Supplier.supplier_name.like(f"%{supplier_name}%"))
        if inspection_item:
            query = query.filter(InspectionTask.inspection_item.like(f"%{inspection_item}%"))
        if laboratory:
            query = query.filter(InspectionTask.laboratory == laboratory)
        if inspector:
            query = query.filter(InspectionTask.inspector.like(f"%{inspector}%"))
        if status:
            query = query.filter(InspectionTask.status == status)
        if result_status:
            query = query.filter(InspectionTask.result_status == result_status)

        start_dt = _parse_date(start_date)
        end_dt = _parse_date(end_date)
        if start_dt:
            query = query.filter(InspectionTask.created_at >= start_dt)
        if end_dt:
            query = query.filter(InspectionTask.created_at < end_dt)

        total = query.count()
        query = query.order_by(InspectionTask.created_at.desc())
        offset = (page - 1) * page_size
        tasks = query.offset(offset).limit(page_size).all()

        results = []
        for t in tasks:
            standard_parts = []
            if t.standard_min is not None and t.standard_max is not None:
                standard_parts.append(f"{t.standard_min}~{t.standard_max}")
            elif t.standard_min is not None:
                standard_parts.append(f"≥{t.standard_min}")
            elif t.standard_max is not None:
                standard_parts.append(f"≤{t.standard_max}")
            if t.standard_tolerance:
                standard_parts.append(f"±{t.standard_tolerance}")

            results.append({
                "task_id": t.id,
                "task_code": t.task_code,
                "sample_code": t.sample.sample_code if t.sample else "",
                "supplier_name": t.sample.supplier.supplier_name if (t.sample and t.sample.supplier) else "",
                "product_name": t.sample.product_name if t.sample else "",
                "sample_type": t.sample.sample_type if t.sample else "",
                "inspection_item": t.inspection_item,
                "laboratory": t.laboratory,
                "inspector": t.inspector,
                "standard": " ".join(standard_parts) if standard_parts else "无",
                "unit": t.unit or "",
                "method": t.method or "",
                "status": t.status,
                "result_value": t.result_value,
                "result_status": t.result_status or "",
                "created_at": t.created_at.strftime("%Y-%m-%d %H:%M:%S") if t.created_at else "",
                "inspected_at": t.inspected_at.strftime("%Y-%m-%d %H:%M:%S") if t.inspected_at else "",
                "inspected_by": t.inspected_by or ""
            })

        return {
            "success": True,
            "total": total,
            "page": page,
            "page_size": page_size,
            "total_pages": (total + page_size - 1) // page_size if page_size > 0 else 0,
            "data": results
        }
    finally:
        db.close()


def query_rectification_orders(
    supplier_name: Optional[str] = None,
    status: Optional[str] = None,
    is_escalated: Optional[bool] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    page: int = 1,
    page_size: int = 50
) -> Dict[str, Any]:
    db = SessionLocal()
    try:
        query = db.query(RectificationOrder).join(Supplier, RectificationOrder.supplier_id == Supplier.id)

        if supplier_name:
            query = query.filter(Supplier.supplier_name.like(f"%{supplier_name}%"))
        if status:
            query = query.filter(RectificationOrder.status == status)
        if is_escalated is not None:
            query = query.filter(RectificationOrder.is_escalated == is_escalated)

        start_dt = _parse_date(start_date)
        end_dt = _parse_date(end_date)
        if start_dt:
            query = query.filter(RectificationOrder.created_at >= start_dt)
        if end_dt:
            query = query.filter(RectificationOrder.created_at < end_dt)

        total = query.count()
        query = query.order_by(RectificationOrder.created_at.desc())
        offset = (page - 1) * page_size
        orders = query.offset(offset).limit(page_size).all()

        results = []
        for o in orders:
            results.append({
                "order_id": o.id,
                "rectify_code": o.rectify_code,
                "supplier_code": o.supplier.supplier_code if o.supplier else "",
                "supplier_name": o.supplier.supplier_name if o.supplier else "",
                "sample_code": o.sample.sample_code if o.sample else "",
                "product_name": o.sample.product_name if o.sample else "",
                "failed_items": o.failed_items,
                "description": o.description or "",
                "deadline": o.deadline.strftime("%Y-%m-%d %H:%M:%S") if o.deadline else "",
                "status": o.status,
                "is_escalated": o.is_escalated,
                "escalated_at": o.escalated_at.strftime("%Y-%m-%d %H:%M:%S") if o.escalated_at else "",
                "rectify_measures": o.rectify_measures or "",
                "rectified_at": o.rectified_at.strftime("%Y-%m-%d %H:%M:%S") if o.rectified_at else "",
                "verified_by": o.verified_by or "",
                "verified_at": o.verified_at.strftime("%Y-%m-%d %H:%M:%S") if o.verified_at else "",
                "created_at": o.created_at.strftime("%Y-%m-%d %H:%M:%S") if o.created_at else ""
            })

        return {
            "success": True,
            "total": total,
            "page": page,
            "page_size": page_size,
            "total_pages": (total + page_size - 1) // page_size if page_size > 0 else 0,
            "data": results
        }
    finally:
        db.close()


def _apply_style_to_ws(ws, df: pd.DataFrame, title: str):
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    header_fill = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.cell(row=1, column=1, value=title).font = Font(size=14, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(df.columns), 1))
    ws["A1"].alignment = center_align
    ws.row_dimensions[1].height = 25

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 3):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            cell.alignment = center_align
            if r_idx == 3:
                cell.fill = header_fill
                cell.font = header_font

    for col_idx, col_name in enumerate(df.columns, 1):
        max_len = max(
            [len(str(col_name))] +
            [len(str(v)) for v in df[col_name].astype(str).tolist()]
        )
        ws.column_dimensions[ws.cell(row=3, column=col_idx).column_letter].width = min(max_len + 4, 50)


def export_to_excel(
    query_type: str = "samples",
    supplier_name: Optional[str] = None,
    inspection_item: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    status: Optional[str] = None,
    extra_filters: Optional[Dict[str, Any]] = None,
    operator: str = "system"
) -> Dict[str, Any]:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{query_type}_export_{timestamp}.xlsx"
    filepath = os.path.join(EXPORT_DIR, filename)

    try:
        wb = Workbook()
        default_ws = wb.active

        if query_type == "samples":
            result = query_samples(
                supplier_name=supplier_name,
                inspection_item=inspection_item,
                start_date=start_date,
                end_date=end_date,
                status=status,
                page=1, page_size=10000
            )
            df = pd.DataFrame(result.get("data", []))
            default_ws.title = "样品记录"
            title = f"样品检测记录导出 ({start_date or '全部'} ~ {end_date or '全部'})"
            _apply_style_to_ws(default_ws, df, title)

            if not df.empty and "inspection_items" in df.columns:
                df = df.drop(columns=["inspection_items"])

        elif query_type == "inspection_tasks":
            result = query_inspection_tasks(
                supplier_name=supplier_name,
                inspection_item=inspection_item,
                start_date=start_date,
                end_date=end_date,
                status=status,
                page=1, page_size=10000
            )
            df = pd.DataFrame(result.get("data", []))
            default_ws.title = "检测任务"
            title = f"检测任务记录导出 ({start_date or '全部'} ~ {end_date or '全部'})"
            _apply_style_to_ws(default_ws, df, title)

        elif query_type == "rectification":
            result = query_rectification_orders(
                supplier_name=supplier_name,
                start_date=start_date,
                end_date=end_date,
                status=status,
                page=1, page_size=10000
            )
            df = pd.DataFrame(result.get("data", []))
            default_ws.title = "整改工单"
            title = f"整改工单记录导出 ({start_date or '全部'} ~ {end_date or '全部'})"
            _apply_style_to_ws(default_ws, df, title)

        elif query_type == "all":
            result_s = query_samples(
                supplier_name=supplier_name, inspection_item=inspection_item,
                start_date=start_date, end_date=end_date, status=status,
                page=1, page_size=10000
            )
            df_s = pd.DataFrame(result_s.get("data", []))
            if "inspection_items" in df_s.columns:
                df_s = df_s.drop(columns=["inspection_items"])
            default_ws.title = "样品记录"
            _apply_style_to_ws(default_ws, df_s, "样品检测记录")

            ws2 = wb.create_sheet("检测任务")
            result_t = query_inspection_tasks(
                supplier_name=supplier_name, inspection_item=inspection_item,
                start_date=start_date, end_date=end_date, status=status,
                page=1, page_size=10000
            )
            df_t = pd.DataFrame(result_t.get("data", []))
            _apply_style_to_ws(ws2, df_t, "检测任务明细")

            ws3 = wb.create_sheet("整改工单")
            result_r = query_rectification_orders(
                supplier_name=supplier_name,
                start_date=start_date, end_date=end_date, status=status,
                page=1, page_size=10000
            )
            df_r = pd.DataFrame(result_r.get("data", []))
            _apply_style_to_ws(ws3, df_r, "整改工单记录")
        else:
            return {"success": False, "error": f"不支持的导出类型: {query_type}"}

        wb.save(filepath)

        log_operation(
            "批量导出",
            f"导出类型:{query_type}，筛选条件:供应商={supplier_name},"
            f"检测项目={inspection_item},时间={start_date}~{end_date}，"
            f"导出文件:{filename}",
            operator=operator
        )
        write_audit_log(
            f"数据导出 | 类型:{query_type} | 文件:{filename} | 操作人:{operator}"
        )

        return {
            "success": True,
            "export_type": query_type,
            "filepath": filepath,
            "filename": filename,
            "record_count": len(df) if query_type != "all" else {
                "samples": len(df_s),
                "tasks": len(df_t),
                "rectifications": len(df_r)
            }
        }
    except Exception as e:
        return {"success": False, "error": str(e)}
